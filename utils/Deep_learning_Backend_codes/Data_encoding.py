import xlwt
from xlwt import Workbook
# Reading an Excel file using Python
import xlrd
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from utils.Deep_learning_Backend_codes.xls_to_csv import convert_xls_to_csv
# Give the location of the file
#loc = ("Remodified_data.csv")

path = "utils/input_data.xlsx"

def run(): 
    # To open the workbook 
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)
    # Get workbook active sheet object 
    # from the active attribute 
    sheet_obj = wb_obj.active 
    
    # To open Workbook
    #wb = xlrd.open_workbook(loc)
    wb = Workbook()
    sheet1 = wb.add_sheet('Sheet1')
    t=0
    Seq=["f","g"]
    # For row 0 and column 0
    #cell_obj = sheet_obj.cell(row = x, column = 0)
    x=2
    count=0
    while (t<=10):
        temp_seq= sheet_obj.cell(row = x, column = 1)
        Seq[0]=temp_seq.value
        temp_seq=sheet_obj.cell(row = x, column = 2)
        Seq[1]=temp_seq.value
        temp_seq=sheet_obj.cell(row = x, column = 3)
        lable=temp_seq.value
        im = np.zeros((10, 23))
        #print(Seq,lable)
        #print(im)
        if Seq[0]==None or Seq[1]==None: break
        for i in range(0,2):
            for j in range(0,23):
                # print("i = ",i,"j= ", j,Seq[i][j])
                if Seq[i][j]=='A':
                    sheet1.write(count+(i*5), j, 1)
                    sheet1.write( count+1+(i*5), j,0)
                    sheet1.write( count+2+(i*5), j,0)
                    sheet1.write( count+3+(i*5), j,0)
                    sheet1.write( count+4+(i*5), j,0)
                    im[i*5][j]=255
                    im[1+(i*5)][j]=0
                    im[2+(i*5)][j]=0
                    im[3+(i*5)][j]=0
                    im[4+(i*5)][j]=0
                elif Seq[i][j]=='G':
                    sheet1.write(count+(i*5), j, 0)
                    sheet1.write( count+1+(i*5), j,1)
                    sheet1.write( count+2+(i*5), j,0)
                    sheet1.write( count+3+(i*5), j,0)
                    sheet1.write( count+4+(i*5), j,0)
                    im[i*5][j]=0
                    im[1+i*5][j]=255
                    im[2+(i*5)][j]=0
                    im[3+(i*5)][j]=0
                    im[4+(i*5)][j]=0
                elif Seq[i][j]=='C':
                    sheet1.write(count+(i*5), j, 0)
                    sheet1.write( count+1+(i*5), j,0)
                    sheet1.write( count+2+(i*5), j,1)
                    sheet1.write( count+3+(i*5), j,0)
                    sheet1.write( count+4+(i*5), j,0)
                    im[i*5][j]=0
                    im[1+(i*5)][j]=0
                    im[2+(i*5)][j]=255
                    im[3+(i*5)][j]=0
                    im[4+(i*5)][j]=0
                elif Seq[i][j]=='T':
                    sheet1.write(count+(i*5), j, 0)
                    sheet1.write( count+1+(i*5), j,0)
                    sheet1.write( count+2+(i*5), j,0)
                    sheet1.write( count+3+(i*5), j,1)
                    sheet1.write( count+4+(i*5), j,0)
                    im[i*5][j]=0
                    im[1+(i*5)][j]=0
                    im[2+(i*5)][j]=0
                    im[3+(i*5)][j]=255
                    im[4+(i*5)][j]=0
                elif Seq[i][j]=='N':
                    sheet1.write(count+(i*5), j, 0)
                    sheet1.write( count+1+(i*5), j,0)
                    sheet1.write( count+2+(i*5), j,0)
                    sheet1.write( count+3+(i*5), j,0)
                    sheet1.write( count+4+(i*5), j,1)
                    im[i*5][j]=0
                    im[1+(i*5)][j]=0
                    im[2+(i*5)][j]=0
                    im[3+(i*5)][j]=0
                    im[4+(i*5)][j]=255
                else :
                    sheet1.write(count+(i*5), j,Seq[i][j])
                    sheet1.write( count+1+(i*5), j,Seq[i][j])
                    sheet1.write( count+2+(i*5), j,Seq[i][j])
                    sheet1.write( count+3+(i*5), j,Seq[i][j])
                    sheet1.write( count+4+(i*5), j,Seq[i][j])
            sheet1.write( count+(i*5), 23,lable)
            sheet1.write( count+1+(i*5), 23,lable)
            sheet1.write( count+2+(i*5), 23,lable)
            sheet1.write( count+3+(i*5), 23,lable)
            sheet1.write( count+4+(i*5), 23,lable)
        # plt.imshow(im,cmap='Greys')
        # filename="image"+str(x-1)+".jpg"
        # plt.savefig(filename)
        count=count+10
        x=x+1
    wb.save('utils/Encoded_input_data.xls')

    input_file = 'utils/Encoded_input_data.xls'
    output_file = 'utils/Encoded_input_data.csv'

    convert_xls_to_csv(input_file, output_file)